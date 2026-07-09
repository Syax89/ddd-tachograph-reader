"""Export of parsed tachograph data to Excel, CSV and PDF.

All exports share the formatting in :mod:`core.report_format` (readable
timestamps, humanised column names, nested structures rendered as text), so
the three formats present the same content consistently.
"""
import logging
import re

from core.utils.report_format import records_to_table, section_tables, summary_rows

_log = logging.getLogger("export")

_EXCEL_MAX_ROWS = 50000
_PDF_MAX_ROWS = 1500

_SHEET_NAME_RE = re.compile(r"[\[\]*?:/\\]")


class ExportManager:

    # ── Excel ────────────────────────────────────────────────────────────

    @staticmethod
    def export_to_excel(data, filepath):
        """Multi-sheet .xlsx: Summary, signature details (VU), one sheet per
        data section. Styled headers, row stripes, auto-filter, frozen panes;
        sections are truncated at _EXCEL_MAX_ROWS."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        HEADER_FONT = Font(bold=True, color="FFFFFF")
        HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
        TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
        STRIPE_FILL = PatternFill("solid", fgColor="EFF4FA")

        wb = Workbook()

        def _autosize(ws, headers, rows):
            for idx, header in enumerate(headers, start=1):
                width = len(str(header))
                for row in rows[:200]:
                    if idx <= len(row):
                        width = max(width, len(str(row[idx - 1])))
                ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 9), 55)

        def _write_table(ws, headers, rows, start_row=1):
            for c, header in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=c, value=header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(vertical="center")
            for r, row in enumerate(rows, start=start_row + 1):
                for c, value in enumerate(row, start=1):
                    cell = ws.cell(row=r, column=c, value=value)
                    if (r - start_row) % 2 == 0:
                        cell.fill = STRIPE_FILL
            ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
            if rows:
                last_col = get_column_letter(len(headers))
                ws.auto_filter.ref = f"A{start_row}:{last_col}{start_row + len(rows)}"
            _autosize(ws, headers, rows)

        # Summary sheet
        ws = wb.active
        ws.title = "Summary"
        ws.cell(row=1, column=1, value="DDD Tachograph Report").font = TITLE_FONT
        row = 3
        for field, value in summary_rows(data):
            if field or value:
                ws.cell(row=row, column=1, value=field).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
            row += 1
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 70

        # Signature details (VU)
        used_names = {"Summary"}
        sv = data.get("signature_verification") or {}
        treps = sv.get("treps") or []
        if treps:
            headers, rows = records_to_table(treps)
            trep_sheet = "TREP Signatures"[:31]
            wsx = wb.create_sheet(trep_sheet)
            used_names.add(trep_sheet)
            _write_table(wsx, headers, rows)
        certs = data.get("vu_certificates") or []
        if certs:
            headers, rows = records_to_table(certs)
            cert_sheet = "VU Certificates"[:31]
            wsx = wb.create_sheet(cert_sheet)
            used_names.add(cert_sheet)
            _write_table(wsx, headers, rows)

        # Data sections
        for label, headers, rows, truncated in section_tables(data, max_rows=_EXCEL_MAX_ROWS):
            if truncated:
                _log.warning("Truncating '%s' to %d rows (Excel limit)", label, _EXCEL_MAX_ROWS)
            safe_label = _SHEET_NAME_RE.sub("_", label)[:31].strip()
            if safe_label in used_names:
                idx = 2
                base = safe_label[:28].rstrip("_")
                while f"{base}_{idx}" in used_names:
                    idx += 1
                safe_label = f"{base}_{idx}"[:31]
            used_names.add(safe_label)
            wsx = wb.create_sheet(safe_label)
            _write_table(wsx, headers, rows)

        wb.save(filepath)

    # ── CSV ──────────────────────────────────────────────────────────────

    @staticmethod
    def export_to_csv(data, filepath):
        """Section-block CSV: every section gets a title row, its own header
        and rows, separated by a blank line — readable in any spreadsheet."""
        import csv

        with open(filepath, "w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle, delimiter=";")

            writer.writerow(["DDD TACHOGRAPH REPORT"])
            for field, value in summary_rows(data):
                if field or value:
                    writer.writerow([field, value])
            writer.writerow([])

            wrote_any = False
            for label, headers, rows, truncated in section_tables(data):
                writer.writerow([f"=== {label.upper()} ==="])
                writer.writerow(headers)
                writer.writerows(rows)
                if truncated:
                    writer.writerow(["… (truncated)"])
                writer.writerow([])
                wrote_any = True

            if not wrote_any:
                writer.writerow(["No data found"])

    # ── PDF ──────────────────────────────────────────────────────────────

    @staticmethod
    def export_to_pdf(data, filepath):
        """Professional A4 report (reportlab): cover page with stats, monthly
        activity tables with bold totals and month separators, then one table
        per data section. Truncated at _PDF_MAX_ROWS."""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            LongTable, PageBreak, Paragraph, SimpleDocTemplate, Spacer,
            TableStyle, HRFlowable,
        )

        def _t2m(ts):
            parts = str(ts).split(":")
            if len(parts) == 2:
                return int(parts[0]) * 60 + int(parts[1])
            return 0

        PRIMARY = colors.HexColor("#1F4E79")
        STRIPE = colors.HexColor("#EFF4FA")
        GRID = colors.HexColor("#B8C7D9")
        TOTAL_BG = colors.HexColor("#D6E4F0")
        LIGHT_GRAY = colors.HexColor("#888888")

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("TachoTitle", parent=styles["Title"],
                                     textColor=PRIMARY, fontSize=22, spaceAfter=4,
                                     leading=26)
        subtitle_style = ParagraphStyle("TachoSubtitle", parent=styles["Normal"],
                                        fontSize=9, textColor=LIGHT_GRAY,
                                        spaceAfter=10)
        section_style = ParagraphStyle("TachoSection", parent=styles["Heading2"],
                                       textColor=PRIMARY, fontSize=13,
                                       spaceBefore=16, spaceAfter=6, leading=16)
        cell_style = ParagraphStyle("TachoCell", parent=styles["BodyText"],
                                    fontSize=8, leading=10)
        head_style = ParagraphStyle("TachoHead", parent=cell_style,
                                    textColor=colors.white, fontName="Helvetica-Bold",
                                    fontSize=8, leading=10)
        total_style = ParagraphStyle("TachoTotal", parent=cell_style,
                                     fontName="Helvetica-Bold", fontSize=8)
        note_style = ParagraphStyle("TachoNote", parent=styles["BodyText"],
                                     fontSize=7, textColor=colors.grey)

        page_w, page_h = landscape(A4)
        doc = SimpleDocTemplate(
            filepath, pagesize=(page_w, page_h),
            leftMargin=15 * mm, rightMargin=15 * mm,
            topMargin=15 * mm, bottomMargin=15 * mm,
            title="DDD Tachograph Report",
        )
        avail_width = page_w - doc.leftMargin - doc.rightMargin

        def _compute_col_widths(headers, rows, max_chars_per_col=40):
            ncols = len(headers)
            if ncols == 0:
                return []
            # Estimate character widths per column
            widths = [len(str(h)) * 0.55 for h in headers]
            for row in rows[:300]:
                for i, v in enumerate(row):
                    if i >= ncols:
                        break
                    w = len(str(v)) * 0.55
                    widths[i] = max(widths[i], min(w, max_chars_per_col * 0.55))
            total = sum(widths)
            if total == 0:
                return [avail_width / ncols] * ncols
            # Scale to available width, with minimum 25pt per column
            scale = avail_width / total
            return [max(w * scale, 25) for w in widths]

        def _is_total_row(row):
            return row and isinstance(row[0], str) and row[0].endswith(" TOTAL")

        def _is_desc_row(row):
            return row and isinstance(row[0], str) and "UTC" in str(row[0])

        def _is_month_boundary(rows, idx):
            if idx == 0:
                return False
            prev = str(rows[idx - 1][0]) if rows[idx - 1] else ""
            curr = str(rows[idx][0]) if rows[idx] else ""
            # Detect month change: e.g. "31/01/2025" → "01/02/2025"
            if "/" in prev and "/" in curr and len(prev) >= 10 and len(curr) >= 10:
                return prev[3:10] != curr[3:10]
            return False

        def _table(headers, rows, is_activity=False):
            col_widths = _compute_col_widths(headers, rows)
            if not col_widths:
                return Spacer(1, 2)

            # Build table data with per-cell styles
            table_data = [[Paragraph(str(h), head_style) for h in headers]]
            alignments = []
            if is_activity:
                alignments = [TA_LEFT, TA_RIGHT, TA_RIGHT, TA_RIGHT, TA_RIGHT, TA_RIGHT, TA_RIGHT]

            for _idx, row in enumerate(rows):
                cells = []
                for _c, val in enumerate(row):
                    s = str(val) if val else ""
                    st = total_style if _is_total_row(row) else cell_style
                    p = Paragraph(s, st) if s else ""
                    cells.append(p)
                table_data.append(cells)

            t = LongTable(table_data, colWidths=col_widths, repeatRows=1)
            # Build styles
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
                ("GRID", (0, 0), (-1, -1), 0.3, GRID),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
            # Zebra striping for data rows
            for i in range(1, len(table_data)):
                if _is_total_row(rows[i - 1]):
                    style_cmds.append(("BACKGROUND", (0, i), (-1, i), TOTAL_BG))
                elif (i - 1) % 2 == 0:
                    style_cmds.append(("BACKGROUND", (0, i), (-1, i), STRIPE))
                else:
                    style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.white))
            # Month boundary lines
            if is_activity:
                for i in range(1, len(rows)):
                    if _is_month_boundary(rows, i - 1):
                        style_cmds.append(("LINEABOVE", (0, i), (-1, i), 1.2, PRIMARY))
            # Right alignment for numeric columns in activity table
            if is_activity and alignments:
                for c, al in enumerate(alignments):
                    if al == TA_RIGHT:
                        style_cmds.append(("ALIGN", (c, 0), (c, -1), "RIGHT"))
            t.setStyle(TableStyle(style_cmds))
            return t

        story = []

        # ═══════ COVER ═══════
        meta = data.get("metadata", {})
        driver = data.get("driver", {})
        activities = data.get("activities") or []
        events = data.get("events") or []

        story.append(Paragraph("Tachograph Report", title_style))
        story.append(Paragraph(
            f"{meta.get('filename', '')}  ·  {meta.get('generation', '')}  ·  "
            f"{'Vehicle Unit' if meta.get('is_vu') else 'Driver Card'}",
            subtitle_style))
        story.append(HRFlowable(width="100%", thickness=1, color=PRIMARY,
                                spaceBefore=4, spaceAfter=10))

        # Stats cards (driving hours, days, events)
        total_drive = 0
        total_work = 0
        total_rest = 0
        for day in activities:
            if not isinstance(day, dict):
                continue
            changes = day.get("changes") or []
            if not isinstance(changes, list) or len(changes) < 2:
                continue
            for i, ch in enumerate(changes):
                if not isinstance(ch, dict):
                    continue
                act = str(ch.get("activity", "")).upper()
                if act == "DRIVE":
                    t1 = _t2m(str(ch.get("time", "00:00")))
                    t2 = _t2m(str(changes[i + 1].get("time", "00:00"))) if i + 1 < len(changes) else 1440
                    if t2 < t1:
                        t2 += 1440
                    total_drive += t2 - t1
                elif act == "WORK":
                    t1 = _t2m(str(ch.get("time", "00:00")))
                    t2 = _t2m(str(changes[i + 1].get("time", "00:00"))) if i + 1 < len(changes) else 1440
                    if t2 < t1:
                        t2 += 1440
                    total_work += t2 - t1
                elif act == "REST":
                    t1 = _t2m(str(ch.get("time", "00:00")))
                    t2 = _t2m(str(changes[i + 1].get("time", "00:00"))) if i + 1 < len(changes) else 1440
                    if t2 < t1:
                        t2 += 1440
                    total_rest += t2 - t1

        stats_data = [
            ["Drive", "Work", "Rest", "Active days", "Events"],
            [f"{total_drive // 60}h {total_drive % 60}m",
             f"{total_work // 60}h {total_work % 60}m",
             f"{total_rest // 60}h {total_rest % 60}m",
             str(len([d for d in activities if isinstance(d, dict) and d.get("changes")])),
             str(len(events)),
            ],
        ]
        stat_table = LongTable(stats_data, colWidths=[avail_width / 5] * 5)
        stat_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 1), (-1, 1), 12),
            ("TEXTCOLOR", (0, 1), (-1, 1), PRIMARY),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("GRID", (0, 0), (-1, -1), 0.3, GRID),
        ]))
        story.append(stat_table)
        story.append(Spacer(1, 8 * mm))

        # Driver info
        if driver.get("surname"):
            name = f"{driver.get('firstname', '')} {driver.get('surname', '')}".strip()
            driver_info = [
                f"Driver: {name}",
                f"Card: {driver.get('card_number', 'N/A')}",
                f"Expiry: {driver.get('expiry_date', 'N/A')}",
            ]
            story.append(Paragraph("  ·  ".join(driver_info), note_style))
            story.append(Spacer(1, 6 * mm))

        story.append(PageBreak())

        # Signature details (VU)
        sv = data.get("signature_verification") or {}
        treps = sv.get("treps") or []
        if treps:
            story.append(Paragraph("Signature Verification", section_style))
            headers, rows = records_to_table(treps)
            story.append(_table(headers, rows))
        certs = data.get("vu_certificates") or []
        if certs:
            story.append(Paragraph("VU Certificates (CVC)", section_style))
            headers, rows = records_to_table(certs)
            story.append(_table(headers, rows))

        # ═══════ DATA SECTIONS ═══════
        first_section = True
        for label, headers, rows, truncated in section_tables(data, max_rows=_PDF_MAX_ROWS):
            if first_section:
                story.append(PageBreak())
                first_section = False
            is_act = (label == "Daily Activities")

            if is_act:
                # Split by month: each month gets its own table + page break.
                months = []
                current_month = []
                for row in rows:
                    if _is_total_row(row):
                        current_month.append(row)
                        months.append(current_month)
                        current_month = []
                    else:
                        current_month.append(row)
                if current_month and not _is_desc_row(current_month[0]):
                    months.append(current_month)

                for mi, month_rows in enumerate(months):
                    if mi > 0:
                        story.append(PageBreak())
                    # Extract month name from the TOTAL row (last row)
                    month_name = ""
                    for r in month_rows:
                        if _is_total_row(r):
                            month_name = str(r[0]).replace(" TOTAL", "")
                            break
                    story.append(Paragraph(
                        f"Daily Activities — {month_name} ({len(month_rows)} rows)",
                        section_style))
                    # Prepend header and description for first month
                    display_headers = headers
                    display_rows = list(month_rows)
                    story.append(_table(display_headers, display_rows, is_activity=True))
                    story.append(Spacer(1, 5 * mm))
            else:
                story.append(Paragraph(
                    f"{label} ({len(rows)}{'+' * truncated})", section_style))
                story.append(_table(headers, rows))
                if truncated:
                    story.append(Paragraph(
                        f"Truncated to {_PDF_MAX_ROWS} rows — use Excel/JSON for full data.",
                        note_style))
                story.append(Spacer(1, 5 * mm))

        def _footer(canvas, document):
            canvas.saveState()
            canvas.setFont("Helvetica", 7)
            canvas.setFillColor(colors.grey)
            canvas.drawString(doc.leftMargin, 8 * mm,
                              f"{meta.get('filename', '')}  ·  "
                              f"TachoReader v{__import__('core.utils.version', fromlist=['__version__']).__version__}")
            canvas.drawRightString(page_w - doc.rightMargin, 8 * mm,
                                   f"Page {document.page}")
            canvas.restoreState()

        doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
