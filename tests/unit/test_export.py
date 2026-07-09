import csv
import os
import sys
import tempfile
import unittest
from copy import deepcopy
from unittest.mock import patch

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from app.export import ExportManager
from core.utils.report_format import build_monthly_activity_report


MOCK_DATA = {
    "metadata": {
        "filename": "mock_file.ddd",
        "parsed_at": "2026-06-08 19:00",
        "integrity_check": "VERIFIED (LOCAL CHAIN)",
        "generation": "G2 (Smart)",
        "file_size_bytes": 12345,
        "coverage_pct": 100.0,
        "is_vu": False,
        "decoder_failure_count": 0,
    },
    "driver": {
        "firstname": "Mario",
        "surname": "Rossi",
        "card_number": "IT123456789"
    },
    "vehicle": {
        "plate": "AA123BB",
        "vin": "VIN1234567890"
    },
    "activities": [
        {
            "date": "01/06/2026",
            "odometer_km": 150,
            "changes": [
                {"activity": "DRIVE", "time": "08:00"},
                {"activity": "REST", "time": "12:00"}
            ]
        },
        {
            "date": "02/06/2026",
            "odometer_km": 200,
            "changes": [
                {"activity": "DRIVE", "time": "09:00"},
                {"activity": "REST", "time": "13:00"}
            ]
        }
    ],
    "events": [
        {"description": "Over speeding", "begin": "2026-06-01T10:30:00+00:00",
         "end": "2026-06-01T10:35:00+00:00", "confidence": "high"},
    ],
    "faults": [
        {"description": "Sensor fault", "begin": "2026-06-02T11:00:00+00:00",
         "end": "N/A", "confidence": "medium"},
    ],
    "locations": [
        {"date": "01/06/2026 10:00", "latitude": 45.4642, "longitude": 9.1900, "description": "Milano"},
        {"date": "02/06/2026 11:00", "latitude": 41.9028, "longitude": 12.4964, "description": "Roma"}
    ],
    "border_crossings": [
        {"confidence": "high", "country_left": "I", "country_entered": "F", "odometer_km": 12345},
    ],
}


class TestExportManager(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.excel_path = os.path.join(self.tmpdir, "export_test.xlsx")
        self.csv_path = os.path.join(self.tmpdir, "export_test.csv")
        self.pdf_path = os.path.join(self.tmpdir, "export_test.pdf")
        self.mock_data = MOCK_DATA

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir)

    def test_export_to_excel(self):
        ExportManager.export_to_excel(self.mock_data, self.excel_path)
        self.assertTrue(os.path.exists(self.excel_path))

        from openpyxl import load_workbook
        wb = load_workbook(self.excel_path)
        sheets = set(wb.sheetnames)

        self.assertIn("Summary", sheets)
        self.assertIn("Daily Activities", sheets)
        self.assertIn("Events", sheets)
        self.assertIn("Faults", sheets)
        self.assertIn("GPS Locations", sheets)
        self.assertIn("Border Crossings", sheets)

        # Summary carries file, driver and vehicle info
        summary_text = "\n".join(
            str(c.value) for row in wb["Summary"].iter_rows() for c in row if c.value)
        self.assertIn("mock_file.ddd", summary_text)
        self.assertIn("Mario Rossi", summary_text)
        self.assertIn("AA123BB", summary_text)

        # Daily Activities: monthly grouped report with hours columns + monthly totals
        ws = wb["Daily Activities"]
        # Description in row 1, headers in row 2
        self.assertIn("Daily activity", str(ws.cell(row=1, column=1).value))
        headers = [c.value for c in ws[2]]
        self.assertIn("Date", headers)
        self.assertIn("Drive (h)", headers)
        self.assertIn("Work (h)", headers)
        self.assertIn("Rest (h)", headers)
        self.assertIn("Available (h)", headers)
        self.assertIn("Unknown (h)", headers)
        self.assertIn("Total (h)", headers)
        rows_read = ws.max_row - 2  # skip desc row and header
        self.assertGreaterEqual(rows_read, 3)  # 2 days + 1 monthly total

        # Events: humanised columns, formatted timestamps, hidden keys dropped
        ws = wb["Events"]
        # Row 1 = description, row 2 = headers, row 3+ = data
        headers = [c.value for c in ws[2]]
        self.assertIn("Description", headers)
        self.assertNotIn("Confidence", headers)
        values = [c.value for c in ws[3]]
        self.assertIn("2026-06-01 10:30", values)
        # Description is in a merged cell row 1
        desc = ws.cell(row=1, column=1).value
        self.assertIsNotNone(desc)

    def test_export_to_csv(self):
        ExportManager.export_to_csv(self.mock_data, self.csv_path)
        self.assertTrue(os.path.exists(self.csv_path))

        with open(self.csv_path, encoding="utf-8-sig") as handle:
            text = handle.read()

        self.assertIn("DDD TACHOGRAPH REPORT", text)
        self.assertIn("Mario Rossi", text)
        self.assertIn("=== DAILY ACTIVITIES ===", text)
        self.assertIn("=== EVENTS ===", text)
        # Section description row
        self.assertIn("Logged events", text)
        # Monthly activity report: Date, Odometer, hours columns + monthly totals
        self.assertIn("Date;Odometer km;Drive (h);Work (h);Rest (h);Available (h);Unknown (h);Total (h)", text)
        self.assertIn("01/06/2026", text)
        self.assertIn("04:00", text)
        self.assertIn("06/2026 TOTAL", text)
        # UTC note
        self.assertIn("UTC", text)
        # Formatted timestamp, not raw ISO
        self.assertIn("2026-06-01 10:30", text)
        self.assertNotIn("2026-06-01T10:30:00+00:00", text)

    def test_export_to_pdf(self):
        ExportManager.export_to_pdf(self.mock_data, self.pdf_path)
        self.assertTrue(os.path.exists(self.pdf_path))
        with open(self.pdf_path, "rb") as handle:
            head = handle.read(5)
        self.assertEqual(head, b"%PDF-")
        self.assertGreater(os.path.getsize(self.pdf_path), 1000)

    def test_spreadsheet_exports_neutralize_formula_text(self):
        data = deepcopy(self.mock_data)
        formulas = ["=SUM(1,1)", "+SUM(1,1)", "-SUM(1,1)", "@SUM(1,1)"]
        data["metadata"]["filename"] = formulas[0]
        data["events"] = [
            {"description": formula, "begin": "2026-06-01T10:30:00+00:00"}
            for formula in formulas
        ]
        data["events"][0]["=untrusted_header"] = "value"
        data["events"][0]["negative_number"] = -42

        ExportManager.export_to_csv(data, self.csv_path)
        with open(self.csv_path, newline="", encoding="utf-8-sig") as handle:
            csv_values = [value for row in csv.reader(handle, delimiter=";") for value in row]

        for formula in formulas:
            self.assertIn("'" + formula, csv_values)
        self.assertIn("'=untrusted Header", csv_values)
        self.assertIn("-42", csv_values)
        self.assertNotIn("'-42", csv_values)
        self.assertIn("2026-06-01 10:30", csv_values)

        ExportManager.export_to_excel(data, self.excel_path)
        from openpyxl import load_workbook
        wb = load_workbook(self.excel_path)
        excel_values = [
            cell.value for sheet in wb.worksheets for row in sheet.iter_rows() for cell in row
            if cell.value is not None
        ]

        for formula in formulas:
            self.assertIn("'" + formula, excel_values)
        self.assertIn("'=untrusted Header", excel_values)
        self.assertIn("-42", excel_values)
        self.assertNotIn("'-42", excel_values)
        self.assertIn("2026-06-01 10:30", excel_values)

    def test_pdf_escapes_dynamic_paragraph_text(self):
        data = deepcopy(self.mock_data)
        data["metadata"]["filename"] = "file & <alter> >.ddd"
        data["metadata"]["generation"] = "G2 <untrusted>"
        data["driver"]["surname"] = "<b>Rossi</b> & Co"
        data["events"][0]["description"] = "Text & <b>must not be markup</b> >"

        from reportlab.platypus import Paragraph as ReportLabParagraph

        paragraph_text = []

        def capture_paragraph(text, *args, **kwargs):
            paragraph_text.append(text)
            return ReportLabParagraph(text, *args, **kwargs)

        with patch("reportlab.platypus.Paragraph", side_effect=capture_paragraph):
            ExportManager.export_to_pdf(data, self.pdf_path)

        self.assertIn("Text &amp; &lt;b&gt;must not be markup&lt;/b&gt; &gt;", paragraph_text)
        self.assertIn(
            "file &amp; &lt;alter&gt; &gt;.ddd  ·  G2 &lt;untrusted&gt;  ·  Driver Card",
            paragraph_text,
        )
        self.assertIn("Mario &lt;b&gt;Rossi&lt;/b&gt; &amp; Co", paragraph_text)

    def test_pdf_handles_malformed_activity_times_without_changing_valid_totals(self):
        _, valid_rows = build_monthly_activity_report(self.mock_data["activities"])
        self.assertEqual(valid_rows[0][2], "04:00")
        self.assertEqual(valid_rows[0][4], "12:00")
        self.assertEqual(valid_rows[0][-1], "16:00")

        malformed_activities = [{
            "date": "03/06/2026",
            "changes": [
                {"activity": "DRIVE", "time": "not-a-time"},
                {"activity": "REST", "time": "12:00"},
            ],
        }]
        _, malformed_rows = build_monthly_activity_report(malformed_activities)
        self.assertEqual(malformed_rows[0][2], "00:00")
        self.assertEqual(malformed_rows[0][4], "12:00")
        self.assertEqual(malformed_rows[0][-1], "12:00")

        data = deepcopy(self.mock_data)
        data["activities"] = malformed_activities
        ExportManager.export_to_pdf(data, self.pdf_path)
        self.assertTrue(os.path.exists(self.pdf_path))


if __name__ == "__main__":
    unittest.main()
