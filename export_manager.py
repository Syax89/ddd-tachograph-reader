"""Export of parsed tachograph data to Excel, CSV and PDF.

All exports share the formatting in :mod:`core.report_format` (readable
timestamps, humanised column names, nested structures rendered as text), so
the three formats present the same content consistently.
"""
import logging

from core.i18n import tr
from core.report_format import records_to_table, section_tables, summary_rows

_log = logging.getLogger("export")

_EXCEL_MAX_ROWS = 50000
_PDF_MAX_ROWS = 1500


class ExportManager:

    # ── Excel ────────────────────────────────────────────────────────────

    @staticmethod
    def export_to_excel(data, filepath):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        HEADER_FONT = Font(bold=True, color="FFFFFF")
        HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
        TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
        STRIPE_FILL = PatternFill("solid", fgColor="EFF4FA")

        wb = Workbook()

        def _autosize(ws, headers, rows):
            for idx, header in enumerate(headers, start=1):
                width = len(str(header))
                for row in rows[:200]:
                    if idx <= len(row):
                        width = max(width, len(str(row[idx - 1])))
                ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 9), 55)

        def _write_table(ws, headers, rows, start_row=1):
            for c, header in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=c, value=header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(vertical="center")
            for r, row in enumerate(rows, start=start_row + 1):
                for c, value in enumerate(row, start=1):
                    cell = ws.cell(row=r, column=c, value=value)
                    if (r - start_row) % 2 == 0:
                        cell.fill = STRIPE_FILL
            ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
            if rows:
                last_col = get_column_letter(len(headers))
                ws.auto_filter.ref = f"A{start_row}:{last_col}{start_row + len(rows)}"
            _autosize(ws, headers, rows)

        # Summary sheet
        ws = wb.active
        ws.title = "Summary"
        ws.cell(row=1, column=1, value="DDD Tachograph Report").font = TITLE_FONT
        row = 3
        for field, value in summary_rows(data):
            if field or value:
                ws.cell(row=row, column=1, value=field).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
            row += 1
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 70

        # Signature details (VU)
        sv = data.get("signature_verification") or {}
        treps = sv.get("treps") or []
        if treps:
            headers, rows = records_to_table(treps)
            wsx = wb.create_sheet("TREP Signatures"[:31])
            _write_table(wsx, headers, rows)
        certs = data.get("vu_certificates") or []
        if certs:
            headers, rows = records_to_table(certs)
            wsx = wb.create_sheet("VU Certificates"[:31])
            _write_table(wsx, headers, rows)

        # Data sections
        for label, headers, rows, truncated in section_tables(data, max_rows=_EXCEL_MAX_ROWS):
            if truncated:
                _log.warning("Truncating '%s' to %d rows (Excel limit)", label, _EXCEL_MAX_ROWS)
            wsx = wb.create_sheet(label[:31])
            _write_table(wsx, headers, rows)

        wb.save(filepath)

    # ── CSV ──────────────────────────────────────────────────────────────

    @staticmethod
    def export_to_csv(data, filepath):
        """Section-block CSV: every section gets a title row, its own header
        and rows, separated by a blank line — readable in any spreadsheet."""
        import csv

        with open(filepath, "w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle, delimiter=";")

            writer.writerow(["DDD TACHOGRAPH REPORT"])
            for field, value in summary_rows(data):
                if field or value:
                    writer.writerow([field, value])
            writer.writerow([])

            wrote_any = False
            for label, headers, rows, truncated in section_tables(data):
                writer.writerow([f"=== {label.upper()} ==="])
                writer.writerow(headers)
                writer.writerows(rows)
                if truncated:
                    writer.writerow(["… (truncated)"])
                writer.writerow([])
                wrote_any = True

            if not wrote_any:
                writer.writerow(["No data found"])

    # ── PDF ──────────────────────────────────────────────────────────────

    @staticmethod
    def export_to_pdf(data, filepath):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            LongTable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, TableStyle,
        )

        PRIMARY = colors.HexColor("#1F4E79")
        STRIPE = colors.HexColor("#EFF4FA")
        GRID = colors.HexColor("#B8C7D9")

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("TachoTitle", parent=styles["Title"],
                                     textColor=PRIMARY, spaceAfter=6)
        section_style = ParagraphStyle("TachoSection", parent=styles["Heading2"],
                                       textColor=PRIMARY, spaceBefore=14, spaceAfter=4)
        cell_style = ParagraphStyle("TachoCell", parent=styles["BodyText"],
                                    fontSize=6.5, leading=8)
        head_style = ParagraphStyle("TachoHead", parent=cell_style,
                                    textColor=colors.white, fontName="Helvetica-Bold")
        note_style = ParagraphStyle("TachoNote", parent=styles["BodyText"],
                                    fontSize=7, textColor=colors.grey)

        page_size = landscape(A4)
        doc = SimpleDocTemplate(
            filepath, pagesize=page_size,
            leftMargin=12 * mm, rightMargin=12 * mm,
            topMargin=12 * mm, bottomMargin=12 * mm,
            title="DDD Tachograph Report",
        )
        avail_width = page_size[0] - doc.leftMargin - doc.rightMargin

        def _table(headers, rows):
            ncols = len(headers)
            col_width = avail_width / ncols
            table_data = [[Paragraph(str(h), head_style) for h in headers]]
            for row in rows:
                table_data.append([Paragraph(str(v), cell_style) if v != "" else ""
                                   for v in row])
            t = LongTable(table_data, colWidths=[col_width] * ncols, repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, STRIPE]),
                ("GRID", (0, 0), (-1, -1), 0.4, GRID),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]))
            return t

        story = [Paragraph("DDD Tachograph Report", title_style)]

        # Summary block as a 2-column table
        sum_rows = [(f, v) for f, v in summary_rows(data) if f or v]
        if sum_rows:
            story.append(_table(["Field", "Value"], sum_rows))

        sv = data.get("signature_verification") or {}
        treps = sv.get("treps") or []
        if treps:
            story.append(Paragraph("TREP Signatures", section_style))
            headers, rows = records_to_table(treps)
            story.append(_table(headers, rows))
        certs = data.get("vu_certificates") or []
        if certs:
            story.append(Paragraph("VU Certificates", section_style))
            headers, rows = records_to_table(certs)
            story.append(_table(headers, rows))

        first_section = True
        for label, headers, rows, truncated in section_tables(data, max_rows=_PDF_MAX_ROWS):
            if first_section:
                story.append(PageBreak())
                first_section = False
            story.append(Paragraph(f"{label} ({len(rows)}{'+' if truncated else ''})",
                                   section_style))
            story.append(_table(headers, rows))
            if truncated:
                story.append(Paragraph(
                    f"Section truncated to {_PDF_MAX_ROWS} rows — use the Excel "
                    f"or JSON export for the complete data.", note_style))
            story.append(Spacer(1, 4 * mm))

        def _footer(canvas, document):
            canvas.saveState()
            canvas.setFont("Helvetica", 7)
            canvas.setFillColor(colors.grey)
            meta = data.get("metadata", {})
            from core.version import APP_NAME, __version__
            canvas.drawString(doc.leftMargin, 7 * mm,
                              f"{meta.get('filename', '')} — {APP_NAME} v{__version__}")
            canvas.drawRightString(page_size[0] - doc.rightMargin, 7 * mm,
                                   f"Page {document.page}")
            canvas.restoreState()

        doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
